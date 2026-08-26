"""Excel export of the comparison.

Part of the deterministic core layer: no AI. This is the "classic" output of
the project - it runs without an API key and produces the file a planner would
actually open on Monday morning.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .reconciler import (
    STATUS_DATE_MISMATCH,
    STATUS_MATCH,
    STATUS_MISSING_ACTUAL,
    STATUS_QTY_MISMATCH,
    STATUS_UNPLANNED,
)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# One fill per status, so the eye finds the problem rows before reading them.
STATUS_FILLS = {
    STATUS_MATCH: PatternFill("solid", fgColor="C6EFCE"),
    STATUS_QTY_MISMATCH: PatternFill("solid", fgColor="FFC7CE"),
    STATUS_DATE_MISMATCH: PatternFill("solid", fgColor="FFEB9C"),
    STATUS_MISSING_ACTUAL: PatternFill("solid", fgColor="F8CBAD"),
    STATUS_UNPLANNED: PatternFill("solid", fgColor="BDD7EE"),
}
STATUS_FONTS = {
    STATUS_MATCH: Font(color="006100"),
    STATUS_QTY_MISMATCH: Font(color="9C0006"),
    STATUS_DATE_MISMATCH: Font(color="9C6500"),
    STATUS_MISSING_ACTUAL: Font(color="833C0C"),
    STATUS_UNPLANNED: Font(color="1F4E79"),
}

SHORTFALL_FONT = Font(color="9C0006", bold=True)
SURPLUS_FONT = Font(color="006100", bold=True)

DATE_COLUMNS = {"planned_date", "actual_date"}
MAX_COLUMN_WIDTH = 42


def write_report(
    comparison: pd.DataFrame, summary: dict, output_path: str | Path
) -> Path:
    """Write the three-sheet Excel report and return the path it was saved to."""
    output_path = Path(output_path)
    if output_path.parent != Path(""):
        output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_summary_sheet(workbook.create_sheet("Summary"), summary)
    _write_table_sheet(
        workbook.create_sheet("Discrepancies"),
        comparison[comparison["status"] != STATUS_MATCH],
    )
    _write_table_sheet(workbook.create_sheet("Full comparison"), comparison)

    workbook.save(output_path)
    return output_path


def _write_summary_sheet(sheet: Worksheet, summary: dict) -> None:
    """Render the aggregate figures as a readable label/value sheet."""
    title = sheet.cell(row=1, column=1, value="Order reconciliation summary")
    title.font = Font(bold=True, size=14)

    rows: list[tuple[str, object]] = [
        ("SECTION", "Coverage"),
        ("Order lines compared", summary["total_order_lines"]),
        ("Lines matching the plan", summary["matched_lines"]),
        ("Lines with a discrepancy", summary["discrepancy_lines"]),
        ("Discrepancy rate (%)", summary["discrepancy_rate_pct"]),
        ("SECTION", "Status breakdown"),
        *[(status, count) for status, count in summary["status_counts"].items()],
        ("SECTION", "Quantities"),
        ("Planned units", summary["total_planned_qty"]),
        ("Actual units", summary["total_actual_qty"]),
        ("Net difference (units)", summary["net_qty_diff"]),
        ("Under-delivered units", summary["under_delivered_qty"]),
        ("Over-delivered units", summary["over_delivered_qty"]),
        ("SECTION", "Scope"),
        ("Date range", f"{summary['date_range']['from']} to {summary['date_range']['to']}"),
        ("Customers", ", ".join(summary["customers"])),
        ("Customers with discrepancies", ", ".join(summary["customers_with_discrepancies"])),
    ]

    if summary["data_quality_flags"]:
        rows.append(("SECTION", "Data quality"))
        rows.extend(summary["data_quality_flags"].items())

    row_number = 3
    for label, value in rows:
        if label == "SECTION":
            cell = sheet.cell(row=row_number, column=1, value=value)
            cell.font = Font(bold=True, color="1F3864")
            row_number += 1
            continue

        sheet.cell(row=row_number, column=1, value=label)
        sheet.cell(row=row_number, column=2, value=value).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        row_number += 1

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 60


def _write_table_sheet(sheet: Worksheet, frame: pd.DataFrame) -> None:
    """Render a comparison DataFrame as a formatted, filterable table."""
    columns = list(frame.columns)

    for column_index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    status_index = columns.index("status") + 1 if "status" in columns else None
    qty_diff_index = columns.index("qty_diff") + 1 if "qty_diff" in columns else None

    for row_offset, (_, record) in enumerate(frame.iterrows(), start=2):
        for column_index, name in enumerate(columns, start=1):
            cell = sheet.cell(
                row=row_offset, column=column_index, value=_excel_value(record[name])
            )
            if name in DATE_COLUMNS:
                cell.number_format = "yyyy-mm-dd"

        if status_index is not None:
            status = record["status"]
            cell = sheet.cell(row=row_offset, column=status_index)
            cell.fill = STATUS_FILLS.get(status, PatternFill())
            cell.font = STATUS_FONTS.get(status, Font())

        if qty_diff_index is not None and pd.notna(record["qty_diff"]):
            cell = sheet.cell(row=row_offset, column=qty_diff_index)
            if record["qty_diff"] < 0:
                cell.font = SHORTFALL_FONT
            elif record["qty_diff"] > 0:
                cell.font = SURPLUS_FONT

    sheet.freeze_panes = "A2"
    if len(frame) > 0:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(frame) + 1}"

    _autofit_columns(sheet, frame, columns)


def _autofit_columns(sheet: Worksheet, frame: pd.DataFrame, columns: list[str]) -> None:
    """Size every column to its widest value, capped so one long cell cannot
    push the rest of the table off the screen."""
    for column_index, name in enumerate(columns, start=1):
        widest_value = frame[name].astype(str).str.len().max() if len(frame) else 0
        width = max(len(name), int(widest_value or 0)) + 2
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            width, MAX_COLUMN_WIDTH
        )


def _excel_value(value: object) -> object:
    """Convert a pandas value into something openpyxl can write.

    pandas' NA/NaT and numpy scalars are not native Python types; without this
    step they land in the file as the literal text "<NA>".
    """
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        return value.item()
    return value
